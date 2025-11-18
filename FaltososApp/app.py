import os
import sys
import uuid
import socket
import threading
import webbrowser
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import pandas as pd

# === Config base ===
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"))
app.secret_key = os.environ.get("SECRET_KEY", "debug_secret")

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# === Leitura simples ===
def read_excel_any(path):
    """Lê Excel (.xls ou .xlsx) a partir da linha 5."""
    engine = "xlrd" if path.lower().endswith(".xls") else "openpyxl"
    print(f"📂 Lendo planilha com engine={engine}")
    df = pd.read_excel(path, engine=engine, header=None, skiprows=4, dtype=object)
    print(f"✅ Planilha lida ({df.shape[0]} linhas x {df.shape[1]} colunas)")
    return df


def clean_dataframe(df):
    """Apenas remove colunas totalmente vazias. Mantém todas as demais (incluindo coluna 0)."""
    df = df.copy()
    antes = df.shape[1]
    df = df.dropna(axis=1, how="all")
    depois = df.shape[1]

    return df.reset_index(drop=True)


def prepare_dataframe(path):
    """Lê e limpa o DataFrame, mantendo a linha 'Falta (Just.)' quando nomes se repetem."""
    df = read_excel_any(path)
    df = clean_dataframe(df)
    print(f"🔍 Total de colunas após limpeza: {df.shape[1]}")

    # --- Remover duplicados na Coluna 1 (índice 1) ---
    if df.shape[1] > 4:
        before = len(df)
        print("🔎 Removendo nomes duplicados na coluna 1, priorizando 'Falta (Just.)' na coluna 4...")

        df["__nome_norm__"] = df.iloc[:, 1].astype(str).str.strip().str.lower()
        df["__falta__"] = df.iloc[:, 3].astype(str).str.strip().str.lower()

        df["__prioridade__"] = df["__falta__"].apply(lambda x: 1 if "falta" in x and "just" in x else 0)

        df = df.sort_values(["__nome_norm__", "__prioridade__"], ascending=[True, False])
        df = df.drop_duplicates(subset="__nome_norm__", keep="first")

        df = df.drop(columns=["__nome_norm__", "__falta__", "__prioridade__"])
        df = df.reset_index(drop=True)

        after = len(df)

    df = df.dropna(how="all")
    df = df.reset_index(drop=True)

    return df


# === Rotas Flask ===
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        f = request.files.get("file")
        if not f:
            flash("Nenhum arquivo enviado.")
            return redirect(request.url)

        uid = str(uuid.uuid4())
        filename = f.filename
        save_path = os.path.join(UPLOAD_FOLDER, f"{uid}_{filename}")
        f.save(save_path)

        try:
            df = prepare_dataframe(save_path)
        except Exception as e:
            flash(f"Erro ao processar planilha: {e}")
            print("❌ Erro:", e)
            return redirect(request.url)

        pickle_path = os.path.join(UPLOAD_FOLDER, f"{uid}.pkl")
        df.to_pickle(pickle_path)
        session.update({
            "pickle_path": pickle_path,
            "idx": 0,
            "orig_path": save_path
        })
        return redirect(url_for("student"))

    return render_template("index.html")


@app.route("/student", methods=["GET", "POST"])
def student():
    if "pickle_path" not in session:
        return redirect(url_for("index"))

    path = session["pickle_path"]
    df = pd.read_pickle(path)
    idx = session.get("idx", 0)

    if idx >= len(df):
        idx = 0

    row = df.iloc[idx]

    nome = str(row.iloc[1]) if len(row) > 1 else ""
    observation = str(row.iloc[6]) if len(row) > 6 else ""

    if request.method == "POST":
        action = request.form.get("action")
        preset = request.form.get("preset")
        obs_text = request.form.get("obs_text", "").strip()

        if preset:
            obs_text = preset

        if obs_text and len(df.columns) > 6:
            df.iat[idx, 6] = obs_text
            df.to_pickle(path)

        if action == "next":
            idx = min(idx + 1, len(df) - 1)
        elif action == "prev":
            idx = max(idx - 1, 0)
        elif action == "download":
            xlsx_path = session["orig_path"].replace(".xls", "_editado.xlsx").replace(".xlsx", "_editado.xlsx")

            df.to_excel(xlsx_path, index=False, header=False, engine="openpyxl")

            # === Estilização sem cabeçalho ===
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            wb = load_workbook(xlsx_path)
            ws = wb.active

            alt_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            font = Font(name="Segoe UI Semibold", size=11, color="000000")

            thin_border = Border(
                left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF")
            )

            alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

            for i, row in enumerate(ws.iter_rows()):
                for cell in row:
                    cell.font = font
                    cell.border = thin_border
                    cell.alignment = alignment
                    cell.fill = alt_fill if i % 2 == 0 else white_fill

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

            wb.save(xlsx_path)

            print(f"💾 Planilha exportada: {xlsx_path}")
            return send_file(xlsx_path, as_attachment=True)

        session["idx"] = idx
        return redirect(url_for("student"))

    return render_template(
        "student.html",
        idx=idx,
        total=len(df),
        student_name=nome,
        observation=observation,
        student_row=row.to_dict()
    )


# === Execução ===
def find_free_port():
    s = socket.socket()
    s.bind(("127.0.0.1", 0))
    port = s.getsockname()[1]
    s.close()
    return port


def run_app():
    port = find_free_port()
    url = f"http://127.0.0.1:{port}/"
    print(f"\n🚀 Servidor rodando em {url}")
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    app.run(port=port, debug=True, use_reloader=False)


if __name__ == "__main__":
    run_app()
